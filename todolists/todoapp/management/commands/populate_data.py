from django.core.management.base import BaseCommand, CommandError
from openpyxl import *
import os
from django.core.wsgi import get_wsgi_application
os.environ['DJANGO_SETTINGS_MODULE'] = 'tolists.settings'
application = get_wsgi_application()
from todoapp import models
from todolists import settings


class Command(BaseCommand):
    help = 'imports the data into the tables'

    def handle(self, *args, **options):
        """to import data into the tables"""

        #####  todolist table #####
        wb = load_workbook("F:\RnD\DjangoProject\\todolists\\todolist.xlsx", read_only=True)
        ws = wb["todolists"]
        table_data = map(lambda row: {'name': row[0].value,'creation_date':row[1].value},ws[2:ws.max_row])

        for row in table_data:

            # print row
            obj = models.Todolists(name=row['name'])
            obj.save()

        #####  Todoietm table  #####
        wb = load_workbook("F:\RnD\DjangoProject\\todolists\\todoitem.xlsx", read_only=True)
        ws = wb['todoitems']
        table_data = map(lambda row: {
            'listname': row[0].value,
            'description': row[1].value,
            'completed': row[2].value,
            'due_by': row[3].value}
                         , ws[2:ws.max_row])

        #print table_data
        for row in table_data:
            #print row
            todolist_obj = models.Todolists.objects.get(name=row['listname'])
            obj = models.Todoitem(description=row['description'], completed=row['completed'],due_by=row['due_by'],
                                 todolists = todolist_obj)
            obj.save()

        print "Imported data successfully"